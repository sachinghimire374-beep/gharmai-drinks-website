#!/usr/bin/env python3
"""Extract the owner's rate list.xlsx into prisma/catalog.json for seeding."""
import json, re
from openpyxl import load_workbook

SRC = "/Users/abiral/Downloads/rate list.xlsx"
OUT = "prisma/catalog.json"

# ---- typo/brand cleanup (professional naming) ----
FIXES = [
    (r"\bplisner\b", "Pilsner"), (r"\bbtl\b", "Bottle"), (r"\bstr\b", "Strong"),
    (r"barah?s+inghe|barasinghe", "Barahsinghe"), (r"\bmrogan\b", "Morgan"),
    (r"chennet", "Chenet"), (r"henesser?y|henessey", "Hennessy"),
    (r"tranqueray", "Tanqueray"), (r"beafeater", "Beefeater"), (r"saphhire", "Sapphire"),
    (r"carbernet", "Cabernet"), (r"chardon+e?y|chardonny", "Chardonnay"),
    (r"shiraj", "Shiraz"), (r"glenl?viit|glenvit|glenl?evit|glenlviit", "Glenlivet"),
    (r"\bmaltn\b", "Malt"), (r"lodon", "London"), (r"jacob\b", "Jacob's"),
    (r"hendrix", "Hendrick's"), (r"jager\s?meister", "Jägermeister"),
    (r"ballentine", "Ballantine's"), (r"jack daniels", "Jack Daniel's"),
    (r"makers mark", "Maker's Mark"), (r"dewars", "Dewar's"), (r"gordons", "Gordon's"),
    (r"absolute", "Absolut"), (r"\bhardys?\b", "Hardys"), (r"manag\b", "Manang"),
    (r"bigmaster", "Big Master"), (r"khukuri", "Khukri"), (r"macallan", "Macallan"),
]

def clean(name: str) -> str:
    s = re.sub(r"\s+", " ", str(name)).strip()
    for pat, rep in FIXES:
        s = re.sub(pat, rep, s, flags=re.I)
    # Title-case words that are all lower, keep known casing
    words = []
    for w in s.split(" "):
        words.append(w if (w[:1].isupper() or "'" in w or w.isupper()) else w.capitalize())
    return " ".join(words)

def size_of(remark):
    if not remark: return ""
    r = str(remark).strip().lower().replace(" ", "")
    if r in ("ltr", "1ltr", "l", "1l"): return "1L"
    m = re.match(r"^([\d.]+)(ml|ltr|l)$", r, re.I)
    if m:
        unit = m.group(2).lower()
        return f"{m.group(1)}{'L' if unit in ('l','ltr') else 'ml'}"
    return str(remark).strip()

# ---- image mapping (reliable Unsplash IDs already proven in this project) ----
IMG = {
    "default":        "https://images.unsplash.com/photo-1569529465841-dfecdab7503b?w=600&h=600&fit=crop",
    "beer":           "https://images.unsplash.com/photo-1535958636474-b021ee887b13?w=600&h=600&fit=crop",
    "beer2":          "https://images.unsplash.com/photo-1608270586620-248524c67de9?w=600&h=600&fit=crop",
    "beer3":          "https://images.unsplash.com/photo-1518176258769-f227c798150e?w=600&h=600&fit=crop",
    "beer4":          "https://images.unsplash.com/photo-1571613316887-6f8d5cbf7ef7?w=600&h=600&fit=crop",
    "corona":         "https://images.unsplash.com/photo-1600788886242-5c96aabe3757?w=600&h=600&fit=crop",
    "whisky":         "https://images.unsplash.com/photo-1569529465841-dfecdab7503b?w=600&h=600&fit=crop",
    "whisky2":        "https://images.unsplash.com/photo-1527281400683-1aae777175f8?w=600&h=600&fit=crop",
    "whisky3":        "https://images.unsplash.com/photo-1585553616435-2dc0a54e271d?w=600&h=600&fit=crop",
    "whisky4":        "https://images.unsplash.com/photo-1615332579037-3c44b3660b53?w=600&h=600&fit=crop",
    "scotch":         "https://images.unsplash.com/photo-1547595628-c61a29f496f0?w=600&h=600&fit=crop",
    "chivas":         "https://images.unsplash.com/photo-1619451334792-150fd785ee74?w=600&h=600&fit=crop",
    "luxewhisky":     "https://images.unsplash.com/photo-1514362545857-3bc16c4c7d1b?w=600&h=600&fit=crop",
    "jameson":        "https://images.unsplash.com/photo-1470337458703-46ad1756a187?w=600&h=600&fit=crop",
    "vodka":          "https://images.unsplash.com/photo-1614963326505-843868e1d83a?w=600&h=600&fit=crop",
    "greygoose":      "https://images.unsplash.com/photo-1607622750671-6cd9a99eabd1?w=600&h=600&fit=crop",
    "vodkafl":        "https://images.unsplash.com/photo-1550985543-f47f38aeee65?w=600&h=600&fit=crop",
    "gin":            "https://images.unsplash.com/photo-1551024709-8f23befc6f87?w=600&h=600&fit=crop",
    "gin2":           "https://images.unsplash.com/photo-1546171753-97d7676e4602?w=600&h=600&fit=crop",
    "rum":            "https://images.unsplash.com/photo-1514218953589-2d7d37efd2dc?w=600&h=600&fit=crop",
    "rum2":           "https://images.unsplash.com/photo-1516997121675-4c2d1684aa3e?w=600&h=600&fit=crop",
    "rum3":           "https://images.unsplash.com/photo-1574096079513-d8259312b785?w=600&h=600&fit=crop",
    "cognac":         "https://images.unsplash.com/photo-1569924995012-c4c706bfcd51?w=600&h=600&fit=crop",
    "winered":        "https://images.unsplash.com/photo-1586370434639-0fe43b2d32e6?w=600&h=600&fit=crop",
    "winewhite":      "https://images.unsplash.com/photo-1553361371-9b22f78e8b1d?w=600&h=600&fit=crop",
    "winerose":       "https://images.unsplash.com/photo-1510812431401-41d2bd2722f3?w=600&h=600&fit=crop",
    "sparkling":      "https://images.unsplash.com/photo-1594372365401-3b5ff14eaaed?w=600&h=600&fit=crop",
}
BRAND_IMG = [
    ("blue label|gold label|green label|double black", "luxewhisky"),
    ("black label|red label|johnnie", "whisky"),
    ("jack daniel", "whisky2"),
    ("glenfiddich|glenlivet|macallan|singleton|monkey shoulder|hibiki", "scotch"),
    ("chivas|ballantine|dewar|jim beam|maker", "chivas"),
    ("jameson|whistler|jägermeister", "jameson"),
    ("hennessy", "cognac"),
    ("grey goose|ciroc", "greygoose"),
    ("absolut peach|absolut lime|absolut vanilla", "vodkafl"),
    ("absolut|smirnoff", "vodka"),
    ("8848|seto bagh|nude|yeti", "vodka"),
    ("hendrick|roku|bottega green", "gin2"),
    ("gin", "gin"),
    ("old monk|khukri|honey hunter", "rum"),
    ("captain morgan|malibu", "rum3"),
    ("corona", "corona"),
    ("carlsberg|tuborg", "beer3"),
    ("gorkha", "beer"),
    ("barahsinghe", "beer2"),
    ("sparkling|moscato", "sparkling"),
    ("rose", "winerose"),
    ("white|chardonnay|sauvignon", "winewhite"),
    ("wine|merlot|cabernet|shiraz|pinot", "winered"),
]
CAT_IMG = {"beer":"beer4","whisky-domestic":"whisky3","whisky-imported":"whisky","single-malt":"scotch",
           "vodka":"vodka","gin":"gin","rum":"rum2","wine":"winered"}

def pick_img(name, cat):
    low = name.lower()
    for pat, key in BRAND_IMG:
        if re.search(pat, low): return IMG[key]
    return IMG[CAT_IMG.get(cat, "default")]

LUX = re.compile(r"blue label|gold label|glenfiddich (15|18|21)|glenlivet (15|18)|macallan|hibiki|hennessy (vsop|xo)|grey goose|ciroc|hendrick|monkey shoulder|maker's|singleton|single barrel|jack daniel's single", re.I)
POP = re.compile(r"red label|jack daniel|tuborg|gorkha strong|khukri|old monk|absolut vodka|jameson|chivas|old durbar yellow|smirnoff", re.I)

def section_to_cat(header):
    h = header.lower()
    if "beer" in h: return "beer"
    if "single malt" in h or "malt" in h: return "single-malt"
    if "imported whisky" in h: return "whisky-imported"
    if "whisky" in h: return "whisky-domestic"
    if "vodka" in h: return "vodka"
    if "gin" in h: return "gin"
    if "rum" in h: return "rum"
    if "wine" in h: return "wine"
    return None

wb = load_workbook(SRC, data_only=True)
ws = wb["Sheet1"]
rows = list(ws.iter_rows(values_only=True))

blocks = [(0,1,2,3), (5,6,7,8), (10,11,12,13), (15,16,17,18)]
items, seen = [], set()
for (c_sn, c_name, c_rate, c_rem) in blocks:
    cat = None
    for r in rows:
        name = r[c_name] if c_name < len(r) else None
        rate = r[c_rate] if c_rate < len(r) else None
        if name and isinstance(name, str):
            maybe = section_to_cat(name)
            low = name.strip().lower()
            if maybe and (rate is None or not isinstance(rate,(int,float))) and low not in ("particulars","particulars "):
                cat = maybe; continue
            if low.startswith(("particulars","s.no")): continue
        if cat and name and isinstance(rate,(int,float)) and rate > 0:
            nm = clean(name)
            size = size_of(r[c_rem] if c_rem < len(r) else None)
            full = f"{nm} {size}".strip()
            key = full.lower()
            if key in seen:  # duplicate row in sheet — keep first
                continue
            seen.add(key)
            items.append({
                "name": full, "baseName": nm, "size": size, "price": int(rate), "cat": cat,
                "image": pick_img(nm, cat),
                "luxury": bool(LUX.search(nm)),
                "badge": "PREMIUM" if LUX.search(nm) else ("POPULAR" if POP.search(nm) else "NONE"),
            })

with open(OUT, "w") as f:
    json.dump(items, f, indent=1)
from collections import Counter
print(f"{len(items)} products →", dict(Counter(i['cat'] for i in items)))
print("luxury:", sum(1 for i in items if i["luxury"]))
